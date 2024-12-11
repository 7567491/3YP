import json
from pathlib import Path
import sys
from typing import List, Dict, Any
from rich.console import Console
from rich.table import Table
from rich.text import Text
import textwrap
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class BlockTableViewer:
    def __init__(self):
        self.console = Console()
        
    def display_blocks(self, json_path: Path) -> None:
        """显示文本块的表格视图并导出到Excel"""
        # 加载数据
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        blocks = data["blocks"]
        total_blocks = data["total_blocks"]
        
        # 创建表格并显示
        self._display_console_table(blocks, total_blocks)
        
        # 导出到Excel
        self._export_to_excel(blocks, total_blocks)
        
    def _display_console_table(self, blocks: List[Dict], total_blocks: int) -> None:
        """在控制台显示表格"""
        table = Table(
            title=f"文本块分析结果 (共 {total_blocks} 块)",
            show_lines=True,
            width=None
        )
        
        # 添加列
        table.add_column("序号", justify="right", style="cyan", no_wrap=True)
        table.add_column("一级标题", style="magenta")
        table.add_column("二级标题", style="blue")
        table.add_column("文本预览", style="green")
        table.add_column("类型", justify="center", style="yellow")
        table.add_column("长度", justify="right", style="cyan")
        table.add_column("页码", justify="right", style="cyan")
        
        # 添加行
        for i, block in enumerate(blocks, 1):
            preview = self._get_text_preview(block["text"])
            table.add_row(
                str(i),
                self._format_title(block["h1_title"]),
                self._format_title(block["h2_title"]),
                preview,
                block["type"],
                str(block["length"]),
                str(block["page"])
            )
        
        # 显示表格
        self.console.print("\n")
        self.console.print(table)
        self.console.print("\n")
        
        # 显示统计信息
        self._display_stats(blocks)
    
    def _export_to_excel(self, blocks: List[Dict], total_blocks: int) -> None:
        """导出数据到Excel文件"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "文本块分析"
        
        # 设置表头
        headers = ["序号", "一级标题", "二级标题", "文本预览", "类型", "长度", "页码"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加数据
        for i, block in enumerate(blocks, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=block["h1_title"])
            ws.cell(row=row, column=3, value=block["h2_title"])
            ws.cell(row=row, column=4, value=self._get_text_preview(block["text"]))
            ws.cell(row=row, column=5, value=block["type"])
            ws.cell(row=row, column=6, value=block["length"])
            ws.cell(row=row, column=7, value=block["page"])
        
        # 设置列宽
        column_widths = [8, 30, 30, 50, 15, 10, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=len(blocks)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 添加统计信息
        stats_row = len(blocks) + 3
        ws.cell(row=stats_row, column=1, value="统计信息").font = Font(bold=True)
        ws.cell(row=stats_row+1, column=1, value="总块数")
        ws.cell(row=stats_row+1, column=2, value=total_blocks)
        ws.cell(row=stats_row+2, column=1, value="平均长度")
        ws.cell(row=stats_row+2, column=2, value=sum(b["length"] for b in blocks)/len(blocks))
        
        # 按类型统计
        type_stats = {}
        for block in blocks:
            block_type = block["type"]
            if block_type not in type_stats:
                type_stats[block_type] = 0
            type_stats[block_type] += 1
        
        stats_row += 4
        ws.cell(row=stats_row, column=1, value="类型统计").font = Font(bold=True)
        for i, (type_name, count) in enumerate(type_stats.items()):
            ws.cell(row=stats_row+1+i, column=1, value=type_name)
            ws.cell(row=stats_row+1+i, column=2, value=count)
        
        # 保存文件
        excel_path = Path(__file__).parent.parent / "data" / "table.xlsx"
        wb.save(str(excel_path))
        self.console.print(f"\n[green]Excel文件已保存到: {excel_path}[/green]\n")
    
    def _format_title(self, title: str) -> str:
        """格式化标题文本"""
        if not title:
            return "---"
        return textwrap.fill(title, width=30)
    
    def _get_text_preview(self, text: str, max_length: int = 50) -> str:
        """获取文本预览"""
        text = text.replace('\n', ' ').strip()
        if len(text) > max_length:
            return text[:max_length] + "..."
        return text
    
    def _display_stats(self, blocks: List[Dict[str, Any]]) -> None:
        """显示统计信息"""
        stats_table = Table(title="文本块统计", show_header=True, header_style="bold magenta")
        stats_table.add_column("类型", style="cyan")
        stats_table.add_column("数量", justify="right", style="green")
        stats_table.add_column("总字数", justify="right", style="green")
        stats_table.add_column("平均长度", justify="right", style="green")
        
        type_stats = {}
        for block in blocks:
            block_type = block["type"]
            if block_type not in type_stats:
                type_stats[block_type] = {"count": 0, "total_length": 0}
            
            type_stats[block_type]["count"] += 1
            type_stats[block_type]["total_length"] += block["length"]
        
        for block_type, stats in type_stats.items():
            avg_length = stats["total_length"] // stats["count"]
            stats_table.add_row(
                block_type,
                str(stats["count"]),
                str(stats["total_length"]),
                str(avg_length)
            )
        
        self.console.print(stats_table)
        self.console.print("\n")

def main():
    # 设置路径
    base_dir = Path(__file__).parent.parent
    json_path = base_dir / "data" / "cut.json"
    
    # 检查文件是否存在
    if not json_path.exists():
        print(f"错误: 找不到文件 {json_path}")
        sys.exit(1)
    
    # 创建查看器并显示
    viewer = BlockTableViewer()
    viewer.display_blocks(json_path)

if __name__ == "__main__":
    main() 